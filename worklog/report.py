#!/usr/bin/env python3
"""Aggregate worklog/sessions.jsonl into a two-tab billing workbook for the
Cage Rentals app.

Tabs:
  1. "Cage Rentals Base"        -- everything <= BASE_THROUGH (May 25 2026
                                   morning). FROZEN. Fixed-price invoice
                                   (LOCKED_BASE_INVOICE = $700). Do NOT
                                   edit this tab going forward.
  2. "Cage Rentals - Iteration" -- post-launch maintenance + iteration on
                                   the base app. Billed hourly at
                                   MAINT_RATE ($150/hr).

PFA Tier 1 (new product work) is tracked in a SEPARATE folder at
/Users/jacobmagna/pfa-tier-one/worklog/ with its own report.py. The two
projects are intentionally architecturally separate -- different codebases,
different databases, different billing rates -- so this report intentionally
does NOT mix them.

Reports two metrics per session and project-wide:
  - Active session time -- wall clock between consecutive events within a
                           session, EXCLUDING any gap > IDLE_THRESHOLD.
  - AI compute time     -- sum of (response_end - prompt_start) per turn,
                           capped at IDLE_THRESHOLD * 4 to drop pathological
                           turns (e.g. laptop closed mid-response).

Usage:
    python3 report.py [--maint-rate 150] [--out worklog_report.xlsx]
"""
import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins

LOG_PATH = Path(__file__).resolve().parent / "sessions.jsonl"
IDLE_THRESHOLD = timedelta(minutes=15)
LA_TZ = ZoneInfo("America/Los_Angeles")

BASE_THROUGH = datetime(2026, 5, 25, 9, 30, 0, tzinfo=LA_TZ)
LOCKED_BASE_INVOICE = 700.00
MAINT_RATE = 150.00
BASE_REFERENCE_RATE = 200.00

TAB_BASE = "Cage Rentals Base"
TAB_MAINT = "Cage Rentals - Iteration"


def parse_log(path: Path):
    events = []
    if not path.exists():
        return events
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                e = json.loads(line)
                dt = datetime.fromisoformat(e["ts"])
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                e["ts_dt"] = dt
                events.append(e)
            except (json.JSONDecodeError, KeyError, ValueError):
                continue
    return events


def session_metrics(events):
    by_session = defaultdict(list)
    for e in events:
        sid = e.get("session_id") or "unknown"
        by_session[sid].append(e)

    out = {}
    for sid, evs in by_session.items():
        evs.sort(key=lambda x: x["ts_dt"])
        start, end = evs[0]["ts_dt"], evs[-1]["ts_dt"]
        active = timedelta(0)
        for a, b in zip(evs, evs[1:]):
            gap = b["ts_dt"] - a["ts_dt"]
            if gap <= IDLE_THRESHOLD:
                active += gap
        ai_total = timedelta(0)
        turns = 0
        pending_start = None
        for e in evs:
            if e["event"] == "prompt_start":
                pending_start = e["ts_dt"]
            elif e["event"] == "response_end" and pending_start is not None:
                delta = e["ts_dt"] - pending_start
                if delta <= IDLE_THRESHOLD * 4:
                    ai_total += delta
                    turns += 1
                pending_start = None
        out[sid] = {
            "start": start, "end": end,
            "active_seconds": active.total_seconds(),
            "ai_seconds": ai_total.total_seconds(),
            "turns": turns,
        }
    return out


def fmt_hms(seconds: float) -> str:
    s = int(seconds)
    h, rem = divmod(s, 3600)
    m, s = divmod(rem, 60)
    return f"{h}:{m:02d}:{s:02d}"


NCOLS = 9
LAST_COL = "I"


def _styles():
    thin = Side(border_style="thin", color="999999")
    return {
        "bold": Font(bold=True),
        "title": Font(bold=True, size=14),
        "subtitle": Font(italic=True, size=9, color="666666"),
        "subsection": Font(bold=True, size=10, color="333333"),
        "subsection_fill": PatternFill("solid", fgColor="E7E6E6"),
        "header_fill": PatternFill("solid", fgColor="DDDDDD"),
        "total_fill": PatternFill("solid", fgColor="FFF2CC"),
        "billable_fill": PatternFill("solid", fgColor="C6EFCE"),
        "box": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center"),
        "right": Alignment(horizontal="right", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
    }


def _section_banner(ws, row, label, fill, font):
    c = ws.cell(row=row, column=1, value=label)
    c.font = font
    c.fill = fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    return row + 1


def render_body(ws, row, metrics, rate, styles, start_index=1,
                fixed_billable_total=None, fixed_label="LOCKED INVOICE TOTAL"):
    sorted_metrics = sorted(metrics.items(), key=lambda x: x[1]["start"])
    total_active = sum(m["active_seconds"] for m in metrics.values())
    total_ai = sum(m["ai_seconds"] for m in metrics.values())
    total_turns = sum(m["turns"] for m in metrics.values())

    row = _section_banner(ws, row, "Sessions",
                          styles["subsection_fill"], styles["subsection"])

    if not metrics:
        c = ws.cell(row=row, column=1, value="(no sessions yet)")
        c.font = Font(italic=True, color="888888")
        c.alignment = styles["center"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        row += 1
    else:
        headers = ["#", "Date", "Start", "End", "Active", "Active (h)",
                   "AI Compute", "AI (h)", "Turns"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = styles["bold"]
            c.fill = styles["header_fill"]
            c.border = styles["box"]
            c.alignment = styles["center"]
        row += 1
        for offset, (_sid, m) in enumerate(sorted_metrics):
            i = start_index + offset
            start_la = m["start"].astimezone(LA_TZ)
            end_la = m["end"].astimezone(LA_TZ)
            same_day = start_la.date() == end_la.date()
            end_str = end_la.strftime("%H:%M:%S") if same_day else end_la.strftime("%m-%d %H:%M:%S")
            values = [
                i, start_la.strftime("%a %m-%d"),
                start_la.strftime("%H:%M:%S"), end_str,
                fmt_hms(m["active_seconds"]),
                round(m["active_seconds"] / 3600, 2),
                fmt_hms(m["ai_seconds"]),
                round(m["ai_seconds"] / 3600, 2),
                m["turns"],
            ]
            for col, v in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=v)
                c.border = styles["box"]
                c.alignment = styles["center"] if col != 2 else styles["left"]
            row += 1

        total_cells = [
            "", "", "", "TOTAL",
            fmt_hms(total_active), round(total_active / 3600, 2),
            fmt_hms(total_ai), round(total_ai / 3600, 2),
            total_turns,
        ]
        for col, v in enumerate(total_cells, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = styles["bold"]
            c.fill = styles["total_fill"]
            c.border = styles["box"]
            c.alignment = styles["center"]
        row += 1
    row += 1

    daily = defaultdict(lambda: {"sessions": 0, "active": 0.0, "ai": 0.0, "turns": 0})
    for m in metrics.values():
        day = m["start"].astimezone(LA_TZ).date()
        daily[day]["sessions"] += 1
        daily[day]["active"] += m["active_seconds"]
        daily[day]["ai"] += m["ai_seconds"]
        daily[day]["turns"] += m["turns"]

    row = _section_banner(ws, row, "Daily totals",
                          styles["subsection_fill"], styles["subsection"])

    if not daily:
        c = ws.cell(row=row, column=1, value="(no days yet)")
        c.font = Font(italic=True, color="888888")
        c.alignment = styles["center"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        row += 1
    else:
        daily_headers = ["Date", "Day", "Sessions", "Active", "Active (h)",
                         "AI Compute", "AI (h)", "Turns", ""]
        for col, h in enumerate(daily_headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = styles["bold"]
            c.fill = styles["header_fill"]
            c.border = styles["box"]
            c.alignment = styles["center"]
        row += 1
        for day in sorted(daily):
            d = daily[day]
            values = [
                day.strftime("%Y-%m-%d"), day.strftime("%A"),
                d["sessions"], fmt_hms(d["active"]),
                round(d["active"] / 3600, 2), fmt_hms(d["ai"]),
                round(d["ai"] / 3600, 2), d["turns"], "",
            ]
            for col, v in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=v)
                c.border = styles["box"]
                c.alignment = styles["center"]
            row += 1
    row += 1

    row = _section_banner(ws, row, "Billable summary",
                          styles["subsection_fill"], styles["subsection"])

    billable_active = round(total_active / 3600 * rate, 2)
    billable_ai = round(total_ai / 3600 * rate, 2)

    summary_rows = [
        ("Sessions", f"{len(metrics)}"),
        ("Turns (prompt -> response)", f"{total_turns}"),
        ("Total active session time",
         f"{fmt_hms(total_active)}  ({total_active/3600:.2f} hrs)"),
        ("Total AI compute time",
         f"{fmt_hms(total_ai)}  ({total_ai/3600:.2f} hrs)"),
    ]
    if fixed_billable_total is None:
        summary_rows.append(("Hourly rate", f"${rate:.2f} / hr"))
    else:
        summary_rows.append(("Hourly rate (reference only)",
                             f"${rate:.2f} / hr -- invoice is fixed-price, not hourly"))
        summary_rows.append(("At reference rate, active basis",
                             f"${billable_active:,.2f}"))
        summary_rows.append(("At reference rate, AI compute basis",
                             f"${billable_ai:,.2f}"))

    for label, val in summary_rows:
        a = ws.cell(row=row, column=1, value=label)
        a.font = styles["bold"]
        a.alignment = styles["left"]
        a.border = styles["box"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        b = ws.cell(row=row, column=5, value=val)
        b.alignment = styles["left"]
        b.border = styles["box"]
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=NCOLS)
        row += 1

    if fixed_billable_total is not None:
        headline_rows = [(fixed_label, f"${fixed_billable_total:,.2f}")]
    else:
        headline_rows = [
            ("BILLABLE (active session basis)", f"${billable_active:,.2f}"),
            ("BILLABLE (AI compute basis)",     f"${billable_ai:,.2f}"),
        ]
    for label, val in headline_rows:
        a = ws.cell(row=row, column=1, value=label)
        a.font = Font(bold=True, size=12)
        a.fill = styles["billable_fill"]
        a.alignment = styles["left"]
        a.border = styles["box"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        b = ws.cell(row=row, column=5, value=val)
        b.font = Font(bold=True, size=12)
        b.fill = styles["billable_fill"]
        b.alignment = styles["right"]
        b.border = styles["box"]
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=NCOLS)
        row += 1

    out_active = fixed_billable_total if fixed_billable_total is not None else billable_active
    out_ai     = fixed_billable_total if fixed_billable_total is not None else billable_ai
    return row, out_active, out_ai


def _apply_print_setup(ws, last_row, footer_text):
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.35, right=0.35, top=0.4, bottom=0.4,
                                  header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True
    ws.print_area = f"A1:{LAST_COL}{last_row}"
    ws.oddFooter.center.text = footer_text + "  -  &P / &N"
    ws.oddFooter.center.size = 9
    ws.oddFooter.center.color = "888888"


def _apply_column_widths(ws):
    widths = {"A": 5, "B": 11, "C": 10, "D": 14, "E": 11,
              "F": 10, "G": 11, "H": 9, "I": 8}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def _write_sheet_header(ws, title, subtitle, styles):
    row = 1
    t = ws.cell(row=row, column=1, value=title)
    t.font = styles["title"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    row += 1
    sub = ws.cell(row=row, column=1, value=subtitle)
    sub.font = styles["subtitle"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    row += 2
    return row


def write_report(base_metrics, maint_metrics, maint_rate, out_path):
    wb = openpyxl.Workbook()
    styles = _styles()
    gen = datetime.now(LA_TZ).strftime("%Y-%m-%d %H:%M %Z")
    base_cutoff_str = BASE_THROUGH.strftime("%Y-%m-%d %H:%M %Z")

    ws_base = wb.active
    ws_base.title = TAB_BASE
    base_title = "Cage Rentals Base -- fixed-price contract  -  FROZEN"
    base_subtitle = (
        f"All work through {base_cutoff_str}. Invoice handed over Mon May 25, "
        f"2026. Locked at ${LOCKED_BASE_INVOICE:,.2f} flat -- this tab is "
        f"closed and should not be edited or added to. Generated {gen}."
    )
    row = _write_sheet_header(ws_base, base_title, base_subtitle, styles)
    row, base_invoice, _ = render_body(
        ws_base, row, base_metrics, rate=BASE_REFERENCE_RATE, styles=styles,
        start_index=1, fixed_billable_total=LOCKED_BASE_INVOICE,
        fixed_label="LOCKED INVOICE  -  Cage Rentals Base",
    )
    _apply_column_widths(ws_base)
    _apply_print_setup(ws_base, row - 1,
                       "Cage Rentals Base - fixed-price - FROZEN")

    ws_maint = wb.create_sheet(TAB_MAINT)
    maint_title = "Cage Rentals - Iteration -- post-launch maintenance & feature work"
    maint_subtitle = (
        f"All work after {base_cutoff_str}. Billed at ${maint_rate:,.2f} / hr. "
        f"PFA Tier 1 product work is tracked separately in the pfa-tier-one "
        f"folder. Generated {gen}.  -  Times in Los Angeles."
    )
    row = _write_sheet_header(ws_maint, maint_title, maint_subtitle, styles)
    maint_start_index = len(base_metrics) + 1
    row, maint_billable_active, maint_billable_ai = render_body(
        ws_maint, row, maint_metrics, rate=maint_rate, styles=styles,
        start_index=maint_start_index, fixed_billable_total=None,
    )
    _apply_column_widths(ws_maint)
    _apply_print_setup(ws_maint, row - 1,
                       f"Cage Rentals - Iteration - ${maint_rate:.0f}/hr")

    wb.active = wb.sheetnames.index(TAB_MAINT)
    wb.save(out_path)
    return base_invoice, maint_billable_active, maint_billable_ai


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--maint-rate", type=float, default=MAINT_RATE,
                    help=f"Hourly rate for Cage Rentals Iteration "
                         f"(default {MAINT_RATE:.0f})")
    ap.add_argument("--out", type=Path,
                    default=Path(__file__).resolve().parent / "worklog_report.xlsx")
    ap.add_argument("--log", type=Path, default=LOG_PATH)
    args = ap.parse_args()

    events = parse_log(args.log)
    if not events:
        print(f"No events found in {args.log}", file=sys.stderr)
        return 1

    base_events = [e for e in events if e["ts_dt"] <= BASE_THROUGH]
    maint_events = [e for e in events if e["ts_dt"] > BASE_THROUGH]
    base_metrics = session_metrics(base_events)
    maint_metrics = session_metrics(maint_events)

    base_invoice, maint_active, maint_ai = write_report(
        base_metrics, maint_metrics, args.maint_rate, args.out,
    )

    def totals(m):
        return (sum(x["active_seconds"] for x in m.values()) / 3600,
                sum(x["ai_seconds"]     for x in m.values()) / 3600)
    ba, bai = totals(base_metrics)
    ma, mai_ = totals(maint_metrics)

    print(f"Wrote {args.out}")
    print(f"  Base cutoff: {BASE_THROUGH.strftime('%Y-%m-%d %H:%M %Z')}")
    print(f"  {TAB_BASE:30s}: {len(base_metrics):2d} sessions  "
          f"active {ba:5.2f} hrs  AI {bai:5.2f} hrs  "
          f"= ${base_invoice:,.2f} (fixed)")
    print(f"  {TAB_MAINT:30s}: {len(maint_metrics):2d} sessions  "
          f"active {ma:5.2f} hrs  AI {mai_:5.2f} hrs  "
          f"@ ${args.maint_rate:.0f}/hr = "
          f"${maint_active:,.2f} (active) / ${maint_ai:,.2f} (AI)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
